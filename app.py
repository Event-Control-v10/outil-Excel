import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from groq import Groq
from audiorecorder import audiorecorder
import io
import copy

# --- CONFIGURATION ---
st.set_page_config(page_title="Assistant Délégué Médical", layout="wide")

# Design personnalisé "Médical"
st.markdown("""
<style>
    .stApp { background-color: #f0f4f8; }
    .main-card { background: white; padding: 25px; border-radius: 15px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); border-left: 5px solid #2ecc71; }
    h1 { color: #2c3e50; font-family: 'Helvetica', sans-serif; }
    .stButton>button { border-radius: 8px; font-weight: bold; }
</style>
""", unsafe_allow_html=True)

# Initialisation de la session pour l'historique (Undo)
if 'history' not in st.session_state:
    st.session_state.history = [] # Stocke les versions successives du DataFrame
if 'current_df' not in st.session_state:
    st.session_state.current_df = None
if 'original_file' not in st.session_state:
    st.session_state.original_file = None

# API Groq
try:
    client = Groq(api_key=st.secrets["GROQ_API_KEY"])
except:
    st.error("Clé API manquante dans les Secrets.")
    st.stop()

# --- FONCTIONS ---

def transcribe(audio_bytes):
    try:
        audio_bio = io.BytesIO()
        audio_bytes.export(audio_bio, format="wav")
        audio_bio.seek(0)
        return client.audio.transcriptions.create(file=("a.wav", audio_bio), model="whisper-large-v3", language="fr").text
    except: return None

def get_ai_logic(instruction, df_sample):
    prompt = f"""
    Tu es un expert Python spécialisé en Pandas.
    Voici les colonnes du fichier : {list(df_sample.columns)}
    Aperçu des données : {df_sample.head(5).to_string()}

    Instruction : "{instruction}"

    Génère UNIQUEMENT le code Python pour modifier le DataFrame 'df'.
    RÈGLES :
    1. Utilise 'df.at' ou 'df.loc' pour être précis.
    2. Ne crée pas de nouvelles colonnes sauf si demandé.
    3. Pas de texte, pas de ```, juste le code.
    4. Le DataFrame s'appelle toujours 'df'.
    """
    res = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0
    ).choices[0].message.content
    return res.replace("```python", "").replace("```", "").strip()

def save_with_design(original_bytes, modified_df):
    """La clé : charger l'original et n'injecter que les valeurs"""
    original_bio = io.BytesIO(original_bytes)
    wb = load_workbook(original_bio)
    ws = wb.active # On prend la feuille active (Sheet1)

    # On détermine à partir de quelle ligne commencent les données 
    # D'après la capture, les en-têtes sont vers la ligne 4, données ligne 5
    # On va faire correspondre le DF avec la grille Excel
    # Attention : L'index de DF commence à 0, Excel à 1.
    
    # Ici, on suppose que les colonnes du DF correspondent exactement à l'ordre Excel
    # et que les données commencent à la ligne 5 (A5, B5...)
    start_row = 5 
    
    for r_idx, row in modified_df.iterrows():
        for c_idx, value in enumerate(row):
            # c_idx + 1 car Excel commence à 1
            ws.cell(row=start_row + r_idx, column=c_idx + 1).value = value

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# --- INTERFACE ---

st.title("👨‍⚕️ Assistant Médical : Automatisation Excel")
st.write("Pour ta grosse ❤️ - Gagne du temps sur tes rapports.")

uploaded = st.file_uploader("Charge ton fichier prescripteurs", type=["xlsx"])

if uploaded:
    if st.session_state.current_df is None:
        # Premier chargement : on saute les 3 premières lignes vides/titres pour lire proprement
        st.session_state.original_file = uploaded.getvalue()
        st.session_state.current_df = pd.read_excel(io.BytesIO(st.session_state.original_file), skiprows=3)

    df = st.session_state.current_df

    with st.container():
        st.markdown('<div class="main-card">', unsafe_allow_html=True)
        st.write("### 🔍 Aperçu du fichier")
        st.dataframe(df.head(10))
        
        col1, col2 = st.columns(2)
        with col1:
            st.write("🎙️ **Parle pour modifier**")
            audio = audiorecorder("Appuyer pour parler", "Stop")
        with col2:
            st.write("⌨️ **Ou écris l'instruction**")
            text_cmd = st.text_input("Ex: Mets 'A+' dans Category pour Dr IDOHOU")

        instruction = ""
        if len(audio) > 0:
            instruction = transcribe(audio)
            if instruction: st.info(f"Entendu : {instruction}")
        elif text_cmd:
            instruction = text_cmd

        if instruction:
            if st.button("🚀 Appliquer la modification"):
                # Sauvegarde pour Undo
                st.session_state.history.append(df.copy())
                
                # Logique IA
                code = get_ai_logic(instruction, df)
                try:
                    exec(code)
                    st.session_state.current_df = df
                    st.success("C'est fait !")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erreur de logique : {e}")

        st.markdown('</div>', unsafe_allow_html=True)

    # Zone d'actions fixes
    st.write("---")
    c_undo, c_save, c_reset = st.columns(3)
    
    with c_undo:
        if st.button("🔙 Annuler (Undo)"):
            if st.session_state.history:
                st.session_state.current_df = st.session_state.history.pop()
                st.rerun()
            else:
                st.warning("Rien à annuler")

    with c_save:
        if st.button("💾 Générer le fichier final"):
            with st.spinner("Conservation du design original..."):
                final_data = save_with_design(st.session_state.original_file, st.session_state.current_df)
                st.download_button("📥 Télécharger l'Excel parfait", final_data, "liste_prescripteurs_maj.xlsx")

    with c_reset:
        if st.button("🗑️ Tout effacer"):
            st.session_state.current_df = None
            st.session_state.history = []
            st.rerun()
